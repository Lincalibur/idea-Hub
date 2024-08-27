import threading
import time
from datetime import datetime, timedelta
import os
from openpyxl import Workbook, load_workbook
import customtkinter as ctk
from collections import defaultdict
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import uuid
import win32gui  # Import the win32gui module

class SecureTaskTimer:
    def __init__(self):
        self.current_app = None
        self.current_window_title = None
        self.start_time = None
        self.session_id = str(uuid.uuid4())
        self.current_date = datetime.now().strftime("%Y-%m-%d")
        self.week_number = datetime.now().isocalendar()[1]
        self.excel_file = os.path.join(os.path.expanduser("~/Desktop"), f'application_usage_log_Week_{self.week_number}.xlsx')
        self.blacklist = [
            "Password Manager", "SecureTaskTimer", "Application Usage Tracker",
            "Downloads", "src", "File Explorer", "Snipping Tool Overlay",
            "New notification", "Snipping Tool", "Outlook", 
            "Outlook Send/Receive Progress", "Skype [1]", "Leave site?", "Skype [2]"
        ]
        self.tracking_active = False
        self.logs = defaultdict(float)  # Store logs as {("App Name", "Window Title"): Total Time}
        self._initialize_excel_file()

    def _initialize_excel_file(self):
        if not os.path.exists(self.excel_file):
            # Create a new workbook if the file does not exist
            workbook = Workbook()
            workbook.save(self.excel_file)
            print(f"Excel file created on the desktop at: {self.excel_file}")

        # Load the workbook and check if the sheet for today exists
        workbook = load_workbook(self.excel_file)
        if self.current_date not in workbook.sheetnames:
            # Create the sheet for today if it does not exist
            sheet = workbook.create_sheet(self.current_date)
            sheet.append(["Application Name", "Window Title", "Start Time", "End Time", "Time Spent (seconds)", "Session ID"])
            workbook.save(self.excel_file)
            print(f"Sheet for {self.current_date} created.")

    def get_active_window(self):
        window = win32gui.GetForegroundWindow()
        window_title = win32gui.GetWindowText(window)
        app_name = self.extract_app_name(window_title)
        return app_name, window_title

    def extract_app_name(self, window_title):
        # Assuming the app name can be derived from the window title
        return window_title.split(" - ")[-1]

    def log_time(self, app_name, window_title, elapsed_time):
        if app_name and app_name not in self.blacklist:
            self.logs[(app_name, window_title)] += elapsed_time

    def save_logs(self):
        try:
            workbook = load_workbook(self.excel_file)
            sheet = workbook[self.current_date]
            for (app_name, window_title), total_time in self.logs.items():
                start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                end_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                sheet.append([app_name, window_title, start_time, end_time, total_time, self.session_id])
            workbook.save(self.excel_file)
        except Exception as e:
            print(f"Error saving logs: {e}")

    def summarize_usage(self):
        try:
            workbook = load_workbook(self.excel_file)
            sheet = workbook[self.current_date]

            summary = defaultdict(lambda: defaultdict(float))  # Nested defaultdict for application names and window titles
            for row in sheet.iter_rows(min_row=2, values_only=True):
                app_name, window_title, _, _, time_spent, session_id = row
                if app_name and time_spent:
                    if session_id == self.session_id:
                        summary[app_name][window_title] += time_spent

            # Clear old data
            sheet.delete_rows(2, sheet.max_row - 1)
            sheet.append(["Summary"])
            sheet.append(["Application Name", "Window Title", "Total Time (hours)"])

            application_summary = defaultdict(float)
            for app_name, windows in summary.items():
                for window_title, total_time in windows.items():
                    time_spent_hours = total_time / 3600  # Convert to hours
                    application_summary[app_name] += time_spent_hours
                    sheet.append([app_name, window_title, time_spent_hours])

            # Write summary data
            sheet.append([])
            sheet.append(["Overall Summary"])
            sheet.append(["Application Name", "Total Time (hours)"])
            for app_name, total_time in application_summary.items():
                sheet.append([app_name, total_time])

            workbook.save(self.excel_file)

            # Display the pie chart
            self.display_pie_chart(application_summary)

        except Exception as e:
            print(f"Error summarizing usage: {e}")

    def display_pie_chart(self, summary):
        # Filter to show only the top 5 entries
        sorted_summary = sorted(summary.items(), key=lambda x: x[1], reverse=True)
        top_summary = dict(sorted_summary[:5])

        if not top_summary:
            print("No data to display in the pie chart.")
            return

        labels = list(top_summary.keys())
        sizes = list(top_summary.values())
        
        fig, ax = plt.subplots()
        ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
        ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle

        # Embed the pie chart in the tkinter window
        chart_frame = ctk.CTkFrame(app)
        chart_frame.pack(pady=10)
        canvas = FigureCanvasTkAgg(fig, master=chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack()

    def run(self):
        self.tracking_active = True
        self.current_app = None
        self.start_time = None

        while self.tracking_active:
            active_app, window_title = self.get_active_window()
            if active_app and active_app not in self.blacklist:
                if active_app != self.current_app or window_title != self.current_window_title:
                    if self.current_app is not None:
                        end_time = datetime.now()
                        elapsed_time = (end_time - self.start_time).total_seconds()
                        self.log_time(self.current_app, self.current_window_title, elapsed_time)
                    self.current_app = active_app
                    self.current_window_title = window_title
                    self.start_time = datetime.now()
                    app_display_var.set(f"Tracking: {active_app} - {window_title}")
            time.sleep(1)

    def stop_tracking(self):
        if self.current_app is not None:
            end_time = datetime.now()
            elapsed_time = (end_time - self.start_time).total_seconds()
            self.log_time(self.current_app, self.current_window_title, elapsed_time)
        self.save_logs()
        app_display_var.set("Tracking stopped.")
        log_display_var.set(f"Logs saved to: {self.excel_file}")

def start_tracking():
    tracker_thread = threading.Thread(target=tracker.run)
    tracker_thread.start()

def stop_tracking():
    tracker.stop_tracking()

def summarize():
    tracker.summarize_usage()

# Initialize the main tracker object
tracker = SecureTaskTimer()

# Set up the customtkinter GUI
ctk.set_appearance_mode("dark")  # Set the appearance mode (light/dark)
ctk.set_default_color_theme("blue")  # Set the default color theme

app = ctk.CTk()  # Create the main window
app.title("Application Usage Tracker")

# Variables for display
app_display_var = ctk.StringVar(value="No application being tracked.")
log_display_var = ctk.StringVar(value="Logs will be saved after tracking is stopped.")

# GUI Elements
start_button = ctk.CTkButton(app, text="Start Tracking", command=start_tracking)
start_button.pack(pady=10)

stop_button = ctk.CTkButton(app, text="Stop Tracking", command=stop_tracking)
stop_button.pack(pady=10)

summarize_button = ctk.CTkButton(app, text="Summarize", command=summarize)
summarize_button.pack(pady=10)

app_display_label = ctk.CTkLabel(app, textvariable=app_display_var)
app_display_label.pack(pady=10)

log_display_label = ctk.CTkLabel(app, textvariable=log_display_var)
log_display_label.pack(pady=10)

# Start the GUI loop
app.mainloop()
